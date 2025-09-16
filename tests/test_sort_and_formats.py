from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook

from fec_formatter.services import XLSXWriterService, OUTPUT_COLUMNS
from fec_formatter.config import AppConfig


def test_sort_and_formats(tmp_path: Path):
    app = AppConfig()
    writer = XLSXWriterService(app.style)
    rows = [
        (["R","C","Addr","Occ/Emp","2024-01-05","100","IMG1"], None),
        (["R","C","Addr","Occ/Emp","2025/02/01 12:00:00","1,000.50","IMG2"], None),
    ]
    out = tmp_path / "out.xlsx"
    writer.write(rows, out)
    wb = load_workbook(out)
    ws = wb.active
    # Header present
    assert ws.cell(row=1, column=1).value == OUTPUT_COLUMNS[0]
    # Amount cell is numeric
    amount_cell = ws.cell(row=2, column=OUTPUT_COLUMNS.index("Contribution Amount") + 1)
    assert isinstance(amount_cell.value, (int, float))

