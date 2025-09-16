from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import datetime
import re
from pathlib import Path
from typing import Iterable, List, Optional, Sequence, Tuple

from .config import StyleConfig

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.worksheet.worksheet import Worksheet


OUTPUT_COLUMNS = [
    "Recipient",
    "Contributor",
    "Contributor Address",
    "Contributor Occupation/Employer",
    "Contribution Date",
    "Contribution Amount",
    "FEC ID",
]


@dataclass
class FECRowBuilder:
    def matches_filters(
        self,
        row: List[str],
        header: List[str],
        names: Sequence[str],
        ids: Sequence[str],
        name_contains: Sequence[str] = (),
    ) -> bool:
        if not names and not ids:
            if not name_contains:
                return True

        def idx(col: str) -> int:
            try:
                return header.index(col)
            except ValueError:
                return -1

        def norm(s: str) -> str:
            # Normalize case and collapse whitespace for robust matching
            return " ".join((s or "").split()).casefold()

        name_idx = idx("contributor_name")
        id_idx = idx("contributor_id")

        name_set = {norm(n) for n in names}
        id_set = {norm(i) for i in ids}
        contains_list = [norm(c) for c in name_contains]

        name_val = norm(row[name_idx]) if name_idx >= 0 and name_idx < len(row) else ""
        id_val = norm(row[id_idx]) if id_idx >= 0 and id_idx < len(row) else ""

        name_match = bool(name_set) and name_val in name_set
        id_match = bool(id_set) and id_val in id_set
        contains_match = bool(contains_list) and any(c in name_val for c in contains_list)

        return name_match or id_match or contains_match

    def build_row(self, header: List[str], row: List[str]) -> List[str]:
        def get(col: str) -> str:
            try:
                i = header.index(col)
            except ValueError:
                return ""
            return row[i] if i < len(row) else ""

        recipient = f"{get('committee_name')} ({get('committee_id')})".strip()

        contributor_name = get('contributor_name')
        contributor_id = get('contributor_id')
        if contributor_id and contributor_id != "C00401224":
            contributor = f"{contributor_name} ({contributor_id})"
        else:
            contributor = contributor_name

        street = ", ".join([p for p in [get('contributor_street_1'), get('contributor_street_2')] if p]).strip(', ')
        city_state = ", ".join([p for p in [get('contributor_city'), get('contributor_state')] if p]).strip(', ')
        address = f"{street}, {city_state} {get('contributor_zip')}".strip()

        employer = get('contributor_employer')
        occupation = get('contributor_occupation')
        if employer and occupation:
            occ_emp = f"{occupation}/{employer}"
        else:
            occ_emp = employer or occupation or ""

        date = get('contribution_receipt_date')
        amount = get('contribution_receipt_amount')
        image_number = get('image_number')

        return [recipient, contributor, address, occ_emp, date, amount, image_number]


@dataclass
class XLSXWriterService:
    style: StyleConfig

    def write(self, rows_with_links: Iterable[Tuple[List[str], Optional[str]]], output_path: Path) -> None:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws: Worksheet = wb.active
        ws.title = "FEC"

        # Base font (applied as we write cells)
        base_font = Font(name=self.style.base_font_name, size=self.style.base_font_size)

        # Header styles
        header_font = Font(name=self.style.base_font_name, size=self.style.base_font_size, bold=True)
        header_fill = PatternFill(fill_type="solid", fgColor=self.style.header_fill_color)

        # Border
        side = Side(style="thin", color=self.style.border_color)
        border = Border(left=side, right=side, top=side, bottom=side)

        # Write header
        ws.append(OUTPUT_COLUMNS)
        for col_idx in range(1, len(OUTPUT_COLUMNS) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border

        # Body rows
        row_idx = 2
        fec_id_col_index = OUTPUT_COLUMNS.index("FEC ID") + 1
        amount_col_index = OUTPUT_COLUMNS.index("Contribution Amount") + 1
        date_col_index = OUTPUT_COLUMNS.index("Contribution Date") + 1
        for row_values, link in rows_with_links:
            ws.append(row_values)
            # Apply base font and borders
            for col_idx in range(1, len(OUTPUT_COLUMNS) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = base_font
                cell.border = border

            # Hyperlink styling
            if link:
                link_cell = ws.cell(row=row_idx, column=fec_id_col_index)
                link_cell.hyperlink = link
                link_cell.font = Font(name=self.style.base_font_name, size=self.style.base_font_size, underline=self.style.hyperlink_underline, color="0000EE")

            # Convert and format amount as numeric
            amount_cell = ws.cell(row=row_idx, column=amount_col_index)
            parsed_amount = _parse_amount(str(amount_cell.value) if amount_cell.value is not None else "")
            if parsed_amount is not None:
                amount_cell.value = parsed_amount
                amount_cell.number_format = self.style.amount_number_format

            # Convert and format date as datetime
            date_cell = ws.cell(row=row_idx, column=date_col_index)
            parsed_date = _parse_date(str(date_cell.value) if date_cell.value is not None else "")
            if parsed_date is not None:
                date_cell.value = parsed_date
                date_cell.number_format = self.style.date_number_format

            row_idx += 1

        wb.save(output_path)


def _parse_amount(raw: str) -> Optional[float]:
    s = raw.strip()
    if not s:
        return None
    # Remove common currency symbols and thousands separators
    for ch in (",", "$"):
        s = s.replace(ch, "")
    try:
        return float(s)
    except ValueError:
        return None


def _parse_date(raw: str) -> Optional[datetime]:
    s = raw.strip()
    if not s:
        return None
    # Normalize common separators and strip timezone/T separators
    s = s.replace("T", " ")
    s = s.replace("Z", "")
    s = s.replace(".000000", "")
    s = s.replace(".000", "")

    # Try common formats with and without times
    formats = [
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d",
        "%Y/%m/%d %H:%M:%S",
        "%m/%d/%Y",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %H:%M",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass

    # Try parsing only the date portion if time present
    parts = s.split()
    if parts:
        date_part = parts[0]
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y"):
            try:
                return datetime.strptime(date_part, fmt)
            except ValueError:
                pass

    # Handle possible day/month order like DD/MM/YYYY
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        a, b, y = m.groups()
        am, ad = int(a), int(b)
        # If the first field exceeds 12, interpret as DD/MM/YYYY
        if am > 12 and 1 <= ad <= 12:
            try:
                return datetime(int(y), ad, am)
            except ValueError:
                return None
        # Otherwise if both <= 12, prefer MM/DD/YYYY (already tried above)
    return None


